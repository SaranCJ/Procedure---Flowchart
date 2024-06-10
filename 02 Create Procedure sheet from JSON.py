import openpyxl
import tkinter as tk
from tkinter import filedialog
import json

# Function to create the initial structure and fill data from JSON
def create_structure_from_json(ws, procedure_name, sequences, start_row, set_number):
    ws[f'A{start_row}'] = 'ProcedureName'
    ws[f'A{start_row+1}'] = 'Model Numbers'
    ws[f'A{start_row+2}'] = 'TotalSequenceCount'
    ws[f'C{start_row}'] = set_number  # Set the set number in the first row of each set
    
    ws[f'B{start_row}'] = procedure_name  # Set the procedure name
    
    # Write sequence numbers in the B column and corresponding step numbers in the D column
    for i in range(25):  # Ensure all 25 sequences are written
        ws[f'B{start_row+3+i}'] = f'Sequence {i+1}'
        if i < len(sequences):
            for key, value in sequences[i].items():
                ws[f'C{start_row+3+i}'] = key
                ws[f'D{start_row+3+i}'] = f"{value},"  # Append a comma to the sequence number

# Ask for JSON file location
root = tk.Tk()
root.withdraw()  # Hide the root window
json_path = filedialog.askopenfilename(filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
                                       title="Select the JSON file with procedures")

# Load the JSON file
with open(json_path, 'r') as file:
    procedures = json.load(file)

# Ask for the existing Excel file location
excel_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                                        title="Select the existing Excel file")

# Load the existing workbook
wb = openpyxl.load_workbook(excel_path)

# Add a new sheet for procedures
ws = wb.create_sheet(title="Procedures")

# Initialize set number
set_number = 1

# Create the initial structure for each procedure in the JSON
for procedure_name, procedure_details in procedures.items():
    sequences = procedure_details['sequence']
    row_offset = (set_number - 1) * 29  # 28 rows + 1 empty row = 29
    create_structure_from_json(ws, procedure_name, sequences, row_offset + 1, set_number)
    set_number += 1

# Save the workbook
wb.save(excel_path)
print(f"File saved to {excel_path}")
