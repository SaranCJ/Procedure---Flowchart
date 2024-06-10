import openpyxl
import json
from tkinter import filedialog
from tkinter import Tk

# Open a Tkinter window (it will be hidden)
root = Tk()
root.withdraw()

# Ask the user to select an Excel file
file_path = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel Files", "*.xlsx")])

if not file_path:
    print("No file selected. Exiting.")
    exit()

# Load the selected Excel workbook
workbook = openpyxl.load_workbook(file_path)

# Get the active sheet
# sheet = workbook.active
# Select the second sheet (index 1)
sheet = workbook.worksheets[0]

# Create a list to store the data
data_list = []

# Keep track of values in column C to check for duplicates
seen_values = set()

# Find the row where "SequenceName" is present in column A
for row in range(1, sheet.max_row + 1):
    if sheet.cell(row=row, column=1).value == "SequenceName":
        # Check if there are at least two more cells in the row
        if row + 2 <= sheet.max_row:
            # Get values from the same row for columns B and C
            value_B = sheet.cell(row=row, column=2).value
            value_C = sheet.cell(row=row, column=3).value

            # Add a comma to each value in column C
            value_C_with_comma = f"{value_C},"

            # Check for duplicates in column C
            if value_C in seen_values:
                print(f"Duplicate value found in column C: {value_C}")
                input("Press Enter to continue...")

            seen_values.add(value_C)

            # Add the values to the data list
            data_list.append({"value_B": value_B, "value_C_with_comma": value_C_with_comma})

# Save the data list as a JSON file
with open('output.json', 'w') as json_file:
    json.dump(data_list, json_file, indent=4)

# Close the workbook
workbook.close()

print("Selected values with commas in column C written to 'output.json'")
